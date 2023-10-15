# Redo the study using 3 years only
import os
import sys
import shutil
import pandas as pd
import numpy as np
import math
from openpyxl.utils.dataframe import dataframe_to_rows
import csv
import pickle
from concurrent.futures import ThreadPoolExecutor as T
from openpyxl import Workbook, load_workbook

import xlrd
import time
import tarfile
import string
from django.conf import settings
from ..core.utils import log_debug
from ..core.models import Debug


class ProcessData:
    def __init__(self, file_name):
        Debug.objects.all().delete()
        self.years = file_name.split("/")[2].split(".")[0]
        self.data_dir = settings.MEDIA_ROOT + '/research/potential'
        self.output_dir = self.data_dir + '/D' + str(self.years)
        if not os.path.exists(self.output_dir):
            os.makedirs(self.output_dir)

        self.gz_file = '/research/potential/D' + str(self.years) + '.gz'
        if os.path.exists(self.gz_file):
            os.remove(self.gz_file)
        self.gz_file_table = '/media/research/potential/D' + str(self.years) + '.gz'
        self.second_time_save = ''

    # Create a file for every folder in the Y_xxxx.xlsx file where xxx= 1985 ... (year)
    def get_files(self, year):
        files = []
        sr = self.data_dir + '/' + str(year) + ".xlsx"
        w = load_workbook(filename=sr, read_only=False)
        sheets = w.sheetnames
        for f in sheets:
            files.append(f)
        w.close()
        return files

    def set_data(self, year, run=1):
        files = []
        sr = self.data_dir + '/' + str(year) + ".xlsx"
        w = load_workbook(filename=sr, read_only=False)
        sheets = w.sheetnames
        for f in sheets:
            files.append(f)
            if run == 1:
                srr = self.output_dir
                srr += '/' + f + "_RRA.xlsx"
                ws1 = w[f]
                wb2 = Workbook()
                # ws2 = wb2.create_sheet("Data")
                ws2 = wb2.active
                ws2.title = "Data"
                mr = ws1.max_row
                mc = ws1.max_column
                # copying the cell values from source
                # excel file to destination excel file
                for i in range(1, mr + 1):
                    for j in range(1, mc + 1):
                        # reading cell value from source excel file
                        c = ws1.cell(row=i, column=j)
                        # writing the read value to destination excel file
                        ws2.cell(row=i, column=j).value = c.value
                        # saving the destination excel file
                try:
                    wb2.save(srr)
                    wb2.close()
                    wb2 = None
                except Exception as err:
                    print("Error: " + str(err))
        w.close()
        return files

    # analyse a variable =>> output file
    # with folder "min-max" with table with four columns "min-n1, max-n1, min-n2, max-n2"
    def process_files_1(self, srr):
        to_save = []
        try:
            df = pd.read_excel(srr, sheet_name="Data", index_col=0, header=1)
            print("1\n", df, "-"*50)
            df = df.apply(pd.to_numeric, errors='coerce')
            print("2\n", df, "-"*50)
            df_min_max = df.iloc[0:2, :]
            print('df_min_max')
            print(df_min_max)
            print('df_min_max')
            df = df.iloc[2:, :]

            df_index = df.index
            df_columns = df.columns
            first_row = df_min_max.iloc[[0]].values[0]
            second_row = df_min_max.iloc[[1]].values[0]
            diff_row = second_row - first_row
            df_n1 = df.apply(lambda row: (row - first_row)/diff_row, axis=1)
            print("10\n", first_row, "\n11\n", second_row, "\n12\n", diff_row, "\n13\n", df_n1)

            to_save.append((df_n1.copy(), srr, 'Normalized1'))
            self.save_to_excel(df_n1, srr, 'Normalized1')
            df_n2 = df_n1.copy()
            df_n2[df_n2 < 0] = 0
            df_n2[df_n2 > 1] = 1
            to_save.append((df_n2.copy(), srr, 'Normalized2'))
            # self.save_to_excel(df_n2, srr, 'Normalized2')
            # print('to_save')
            # print(to_save)
            # print('to_save')
            if len(df_n1.columns) < 2:
                df_n1["max"] = df_n1[df_n1.columns[0]]  # df_n1["Birth Rate"]
                df_n2["max"] = df_n2[df_n2.columns[0]]  # df_n2["Birth Rate"]
                df_1_2 = pd.merge(left=df_n1, right=df_n2, left_index=True, right_index=True)
                df_1_2.columns = ['min-n1', 'max-n1', 'min-n2', 'max-n2']
                to_save.append((df_1_2.copy(), srr, 'min-max'))
                # self.save_to_excel(df_1_2, srr, 'min-max')
                self.save_to_excel_(to_save, srr)
                return 1
            zero_list = {}
            one_list = {}
            for i in df_n1.index:
                n0 = []
                n1 = []
                for num in df_n1.loc[i]:
                    if not pd.isna(num):
                        if num <= 0:
                            n0.append(num)
                        elif num >= 1:
                            n1.append(num)
                if len(n0) > 0:
                    n0.sort()
                    zero_list[i] = n0
                elif len(n1) > 0:
                    n1.sort()
                    one_list[i] = n1

                a = df_n2.values
                a_1 = a.copy()
                a_1.sort(axis=1)
                for n in range(a_1.shape[1] - 1):
                    if n == 0:
                        d = a_1[:, n + 1:n + 2] - a_1[:, 0:1]
                    else:
                        d = np.concatenate((d, (a_1[:, n + 1:n + 2] - a_1[:, 0:1])), axis=1)
            to_save.append((pd.DataFrame(a_1).copy(), srr, 'Left'))
            # self.save_to_excel(pd.DataFrame(a_1), srr, 'Left')
            a_1 = self.clean_rows(a=a_1, d=d, srr=srr, n=0, to_save=to_save)
            a_1 = -1 * a_1
            a_1.sort(axis=1)
            for n in range(a_1.shape[1] - 1):
                if n == 0:
                    d = a_1[:, n + 1:n + 2] - a_1[:, 0:1]
                else:
                    d = np.concatenate((d, (a_1[:, n + 1:n + 2] - a_1[:, 0:1])), axis=1)

            to_save.append((pd.DataFrame(a_1).copy(), srr, 'right'))
            # self.save_to_excel(pd.DataFrame(a_1), srr, 'right')
            a_1 = self.clean_rows(a=a_1, d=d, srr=srr, n=0, to_save=to_save)
            a_1 = -1 * a_1
            a_1.sort(axis=1)
            a_1 = pd.DataFrame(a_1)
            to_save.append((a_1.copy(), srr, 'Right'))
            # self.save_to_excel(a_1, srr, 'Right')
            a_1 = a_1.apply(self.twenty_rule, axis=1)

            to_save.append((a_1.copy(), srr, 'twenty'))
            # self.save_to_excel(a_1, srr, 'twenty')

            a_1 = a_1.apply(self.twentyfive_rule, axis=1)
            a_1 = a_1.values
            a_1.sort(axis=1)
            a_1 = pd.DataFrame(data=a_1, index=list(df_index))
            a_2 = a_1.copy()
            to_save.append((a_2.copy(), srr, 'twenty_five'))
            # self.save_to_excel(a_2, srr, 'twenty_five')
            ff = []
            for j in a_1.index:
                nn = list(a_1.loc[j])
                if min(nn) == 0:
                    nn = [zero_list[j].pop(0) if i == 0 else i for i in nn]
                if max(nn) == 1:
                    nn = [one_list[j].pop() if i == 1 else i for i in nn]
                # nn.insert(0, j)
                ff.append(nn)
            a_1 = pd.DataFrame(ff, index=list(df_index))
            # a_1.columns = df_columns
            to_save.append((a_1.copy(), srr, 'twenty2_n1'))
            # self.save_to_excel(a_1, srr, 'twenty2_n1')
            # print("9003")
            a_1 = a_1.apply(self.min_max_rule, axis=1)
            a_1.columns = ['min-n1', 'max-n1']
            #
            a_2 = a_2.apply(self.min_max_rule, axis=1)
            a_2.columns = ['min-n2', 'max-n2']
            a_1_2 = pd.merge(left=a_1, right=a_2, left_index=True, right_index=True)
            to_save.append((a_1_2.copy(), srr, 'min-max'))
            # self.save_to_excel(a_1_2, srr, 'min-max')
            print("9005")
            self.save_to_excel_(to_save, srr)
            print("9005-1")
        except Exception as e:
            log_debug('process_files_1: 2001 ' + str(e))
            print('process_files_1: 2001' + str(e))

    # collect min-max from all the files of each variable
    def collect_data(self, year, files):
        sr = self.output_dir +'/SR_Potential_' + str(year) + '.xlsx'

        log_debug('collect_data 1004: ')
        log_debug(sr)
        log_debug('collect_data 1004: ')
        print('collect_data 1004: ')
        print(sr)
        print('collect_data 1004: ')

        n = 0
        sss = ""
        for f in files:
            srr = self.output_dir + "/" + f + "_RRA.xlsx"
            if n > 0:
                try:
                    df = pd.read_excel(srr, sheet_name="min-max", index_col=0, header=1)
                    df1_ = df.iloc[:, [0, 1]]
                    df2_ = df.iloc[:, [2, 3]]
                    df_1 = pd.merge(left=df_1, right=df1_, left_index=True, right_index=True)
                    df_2 = pd.merge(left=df_2, right=df2_, left_index=True, right_index=True)
                except Exception as e:
                    print(e)
                sss += ", "
            else:
                try:
                    df = pd.read_excel(srr, sheet_name="min-max", index_col=0, header=1)
                    df_1 = df.iloc[:, [0, 1]]
                    df_2 = df.iloc[:, [2, 3]]
                except Exception as ex:
                    print(ex)
            sss += "'x" + str(n) + "_min', 'x" + str(n) + "_max'"
            n += 1
        sss1 = "df_1.columns = [" + sss + "]"
        sss2 = "df_2.columns = [" + sss + "]"
        try:
            exec(sss1)
            exec(sss2)
        except Exception as ex:
            print(ex)

        # df_1.columns = ['x0_min', 'x0_max', 'x1_min', 'x1_max', 'x2_min', 'x2_max',
        #                 'x3_min', 'x3_max', 'x4_min', 'x4_max', 'x5_min', 'x5_max',
        #                 'x6_min', 'x6_max']
        # df_2.columns = ['x0_min', 'x0_max', 'x1_min', 'x1_max', 'x2_min', 'x2_max',
        #                 'x3_min', 'x3_max', 'x4_min', 'x4_max', 'x5_min', 'x5_max',
        #                 'x6_min', 'x6_max']

        if not os.path.isfile(sr):
            book = Workbook()
            sheet = book.active
            book.save(sr)
            book.close()
            book = None

        to_save = []
        to_save.append((df_1.copy(), sr, 'Data-n1'))
        to_save.append((df_2.copy(), sr, 'Data-n2'))
        self.save_to_excel_(to_save, sr)

        # self.save_to_excel(df_1, sr, 'Data-n1')
        # self.save_to_excel(df_2, sr, 'Data-n2')

        book = load_workbook(sr)
        sheet_ = book["Sheet"]
        book.remove(sheet_)
        book.save(sr)
        book.close()
        book = None
        log_debug('collect_data 1105: ')

    def main(self):
        year = self.years
        files = self.set_data(year=year, run=1)
        for variable in files:
            ssr = self.output_dir + "/" + variable + "_RRA.xlsx"
            try:
                log_debug('main 1002: ' + ssr)
                print('main 1002: ' + ssr)
                self.process_files_1(ssr)
                print('main 1003: ' + ssr)
                log_debug('main 1003: ' + ssr)
            except Exception as e:
                print(e)
        self.collect_data(year=year, files=files)
        self.process_files_3(year=year)

    def process_files_3(self, year):
        log_debug('process_files_3, 1005: ' + str(year))
        #
        files = self.get_files(year)
        n_files = len(files)
        # --
        sr = self.output_dir + '/SR_Potential_' + str(year) + '.xlsx'
        srs = self.output_dir + '/SR_Potential_Summary_' + str(year) + '.xlsx'

        f_ = "Potential"

        w = load_workbook(filename=sr, read_only=False)
        df = pd.read_excel(sr, sheet_name='Data-n2')
        n_cols = len(df.columns)
        #
        # Get the min distance from min and max
        try:
            for n in range(1, n_files):
                s = "df['d" + str(n) + "_mm'] = abs(df['x0_min'] - df['x" + str(n) + "_min'])"
                exec(s)
            for n in range(1, n_files):
                s = "df['d" + str(n) + "_xx'] = abs(df['x0_max'] - df['x" + str(n) + "_max'])"
                exec(s)
            for n in range(1, n_files):
                s = "df['d" + str(n) + "_mx'] = abs(df['x0_min'] - df['x" + str(n) + "_max'])"
                exec(s)
            for n in range(1, n_files):
                s = "df['d" + str(n) + "_xm'] = abs(df['x0_max'] - df['x" + str(n) + "_min'])"
                exec(s)
            # min of every group
            n_cols_ = n_cols + (n_files-1)
            df['SComb_mm'] = df.iloc[:, n_cols:n_cols_].min(axis=1)
            n_cols_1 = n_cols_ + (n_files-1)
            df['SComb_xx'] = df.iloc[:, n_cols_:n_cols_1].min(axis=1)
            n_cols_2 = n_cols_1 + (n_files-1)
            df['SComb_mx'] = df.iloc[:, n_cols_1:n_cols_2].min(axis=1)
            n_cols_3 = n_cols_2 + (n_files-1)
            df['SComb_xm'] = df.iloc[:, n_cols_2:n_cols_3].min(axis=1)
            # ---------------------------------------------
            # how close every variable to the min of all variables?
            for n in range(1, n_files):
                s = "df['s" + str(n) + "_mm'] = df['d" + str(n) + "_mm'] - df['SComb_mm']"
                exec(s)
            for n in range(1, n_files):
                s = "df['s" + str(n) + "_xx'] = df['d" + str(n) + "_xx'] - df['SComb_xx']"
                exec(s)
            for n in range(1, n_files):
                s = "df['s" + str(n) + "_mx'] = df['d" + str(n) + "_mx'] - df['SComb_mx']"
                exec(s)
            for n in range(1, n_files):
                s = "df['s" + str(n) + "_xm'] = df['d" + str(n) + "_xm'] - df['SComb_xm']"
                exec(s)
            wspd = w.create_sheet("Processed_Data", 2)
            for r in dataframe_to_rows(df, index=False, header=1):
                wspd.append(r)
            #
            n_ = 1
            for n in range(0, n_files):
                l_ = string.ascii_uppercase[n_]
                n_ += 1
                s = "w['Data-n1']['" + l_ + "1'] = w['Data-n2']['" + l_ + "1'] = files[" + str(n) + "]+'_min'"
                exec(s)
                l_ = string.ascii_uppercase[n_]
                n_ += 1
                s = "w['Data-n1']['" + l_ + "1'] = w['Data-n2']['" + l_ + "1'] = files[" + str(n) + "]+'_max'"
                exec(s)

            log_debug('process_files_3, 1006: ' + str(year))

            sim_mm = []
            sim_xx = []
            sim_mx = []
            sim_xm = []
            sc_mm = []
            sc_xx = []
            sc_mx = []
            sc_xm = []
            scs = {}
            for zz in ['mm', 'xx', 'xm', 'mx']:
                for n in range(1, n_files):
                    eval("sim_" + zz + ".append(round(1 - df['d" + str(n) + "_" + zz + "'].mean(),5))")
                    eval("sc_" + zz + ".append(round(1 - df['s" + str(n) + "_" + zz + "'].mean(),5))")
                eval("sim_" + zz + ".append(round(1 - df['SComb_" + zz + "'].mean(),5))")

            for zz in ['mm', 'xx', 'xm', 'mx']:
                sum_ = str(eval("round(sum(sc_" + zz + ") - 0.7 * len(sc_"+zz+"), 5)"))
                exec("sum_" + zz + "=" + str(sum_))
                k = eval("[round((x - 0.7)/" + sum_ + ", 5) for x in sc_" + zz + "]")
                exec("scs_"+zz+"=k.copy()")
                scs[zz] = k
            # --------------------------------------
            log_debug('process_files_3, 1007: ' + str(year))

            wspd1 = w.create_sheet("Significance", 3)
            n1 = 1
            for zz in ['mm', 'xx', 'xm', 'mx']:
                n_ = 0
                for n in range(1, n_files):
                    l_ = string.ascii_uppercase[n_]
                    wspd1[l_+str(n1)] = 'sim'+str(n)+'_'+zz
                    n_ += 1
                l_ = string.ascii_uppercase[n_]
                wspd1[l_+str(n1)] = 'SCom_'+zz
                wspd1.append(eval('sim_'+zz))
                n1 += 2

            # wspd1.append(['', '', '', '', '', ''])
            n1 += 2
            for zz in ['mm', 'xx', 'xm', 'mx']:
                n_ = 0
                for n in range(1, n_files):
                    l_ = string.ascii_uppercase[n_]
                    wspd1[l_+str(n1)] = 's'+str(n)+'_'+zz
                    n_ += 1
                l_ = string.ascii_uppercase[n_]
                wspd1[l_+str(n1)] = 'SCom_'+zz
                wspd1.append(eval('sc_'+zz))
                n1 += 2
            for zz in ['mm', 'xx', 'xm', 'mx']:
                wspd1.append(['', '', '', '', '', ''])
                wspd1.append(['sum_' + zz, '', '', '', '', ''])
                wspd1.append([str(eval("sum_" + zz)), '', '', '', '', ''])
                wspd1.append(['a_' + zz + '_factors', '', '', '', '', ''])
                wspd1.append([str(scs[zz]), '', '', '', '', ''])
            # ------
            df_n1 = pd.read_excel(sr, sheet_name='Data-n1')
            c = list(df_n1.columns)
            c[0] = "ticker"
            df_n1.columns = c
            dfp = pd.DataFrame()
            for i in range(1, n_files):
                try:
                    smm = 'x' + str(i) + '_min'
                    sxx = 'x' + str(i) + '_max'
                    if i == 1:
                        dfp['ticker'] = df_n1['ticker']
                        dfp['min_min'] = scs['mm'][i - 1] * df_n1[smm]
                        dfp['min_max'] = scs['mx'][i - 1] * df_n1[sxx]
                        dfp['max_min'] = scs['xm'][i - 1] * df_n1[smm]
                        dfp['max_max'] = scs['xx'][i - 1] * df_n1[sxx]
                    else:
                        dfp['min_min'] = dfp['min_min'] + scs['mm'][i - 1] * df_n1[smm]
                        dfp['min_max'] = dfp['min_max'] + scs['mx'][i - 1] * df_n1[sxx]
                        dfp['max_min'] = dfp['max_min'] + scs['xm'][i - 1] * df_n1[smm]
                        dfp['max_max'] = dfp['max_max'] + scs['xx'][i - 1] * df_n1[sxx]
                except Exception as e:
                    print('Error\n' + str(i) + "\n" + str(year), '\n---\n', e, '\n---')
            df_nn = df_n1.isnull().iloc[:, 0:n_cols]
            df_nn_sum = df_nn.sum(axis=1)
            for ind in df_n1.index:
                if df_nn_sum[ind] > 0:
                    zmin = 0
                    zmax = 0
                    vv = []
                    aa_min_sc_mm_r = []
                    aa_min_sc_xm_r = []
                    for j in range(1, 6):
                        i = j * 2 - 1
                        if df_nn.iloc[ind:(ind + 1), i + 2:i + 3].values[0][0].astype('bool'):
                            zmin += scs['mm'][j - 1]
                            zmax += scs['xm'][j - 1]
                        else:
                            vv.append(df_n1.iloc[ind:(ind + 1), i + 2:i + 3].values[0][0])
                            aa_min_sc_mm_r.append(sc_mm[j - 1])
                            aa_min_sc_xm_r.append(sc_xm[j - 1])
                    if zmin < 0.2:
                        dfp.iloc[ind:(ind + 1), 1:2] = self.get_potential(vv, aa_min_sc_mm_r)
                    if zmax < 0.2:
                        dfp.iloc[ind:(ind + 1), 3:4] = self.get_potential(vv, aa_min_sc_xm_r)
                    zmin = 0
                    zmax = 0
                    vv = []
                    aa_min_sc_mx_r = []
                    aa_min_sc_xx_r = []
                    for j in range(1, 6):
                        i = j * 2
                        if df_nn.iloc[ind:(ind + 1), i + 2:i + 3].values[0][0].astype('bool'):
                            zmin += scs['mx'][j - 1]
                            zmax += scs['xx'][j - 1]
                        else:
                            vv.append(df_n1.iloc[ind:(ind + 1), i + 2:i + 3].values[0][0])
                            aa_min_sc_mx_r.append(sc_mx[j - 1])
                            aa_min_sc_xx_r.append(sc_xx[j - 1])
                    if zmin < 0.2:
                        dfp.iloc[ind:(ind + 1), 2:3] = self.get_potential(vv, aa_min_sc_mx_r)
                    if zmax < 0.2:
                        dfp.iloc[ind:(ind + 1), 4:5] = self.get_potential(vv, aa_min_sc_xx_r)

            dfp['max_value-n1'] = dfp.max(axis=1)
            dfp['min_value-n1'] = dfp.min(axis=1)
            dfp['max_value-n2'] = dfp['max_value-n1']
            dfp['min_value-n2'] = dfp['min_value-n1']
            dfp.loc[dfp['max_value-n2'] < 0, 'max_value-n2'] = 0
            dfp.loc[dfp['max_value-n2'] > 1, 'max_value-n2'] = 1
            dfp.loc[dfp['min_value-n2'] < 0, 'min_value-n2'] = 0
            dfp.loc[dfp['min_value-n2'] > 1, 'min_value-n2'] = 1

            wspd2 = w.create_sheet("Potential", 4)
            for r in dataframe_to_rows(dfp, index=False, header=1):
                wspd2.append(r)

            log_debug('process_files_3, 1008: ' + str(year))
            #
            n_ = 1
            for n in range(n_files):
                l_ = string.ascii_uppercase[n_]
                w['Processed_Data'][l_+str(1)] = files[n]+'_min'
                n_ += 1
                l_ = string.ascii_uppercase[n_]
                w['Processed_Data'][l_+str(1)] = files[n]+'_max'
                n_ += 1
            #
            w.save(sr)

            ws1 = w[f_]
            wb2 = Workbook()
            ws2 = wb2.active
            ws2.title = f_
            mr = ws1.max_row
            mc = ws1.max_column
            for i in range(1, mr + 1):
                for j in range(1, mc + 1):
                    c = ws1.cell(row=i, column=j)
                    ws2.cell(row=i, column=j).value = c.value
            wb2.save(srs)
            wb2.close()
            wb2 = None
            #
            w.close()
            w = None
        except Exception as ex:
            log_debug('Error 1005: ' + str(ex))
            print('error')
            print(ex)

    def compress(self):
        tar = tarfile.open(settings.MEDIA_ROOT + self.gz_file, "w:gz")
        tar.add(self.output_dir, arcname=self.years)
        tar.close()
        try:
            shutil.rmtree(self.output_dir)
        except OSError as e:
            print("Error: %s - %s." % (e.filename, e.strerror))

    def min_max_rule(self, row):
        row_ = row[0:2].copy()
        row_[:] = np.nan
        row_[0] = row.min()
        row_[1] = row.max()
        return row_

    def twentyfive_rule(self, row):
        n_row = row.count()
        row_best = row
        if (n_row > 4) and (row.max()-row.min() > 0.25):
            n = 1
            min = n_row
            for j in range(n+1):
                row_c=row.copy()
                row_c[:j] = np.nan
                row_c[n_row-(n-j):] = np.nan
                if (row_c.max()-row_c.min()) < min :
                    min = row_c.max()-row_c.min()
                    row_best = row_c.copy()
        if row_best.max() - row_best.min() > 0.25:
            row_best[:] = np.nan
        return row_best

    def twenty_rule(self, row):
        n_row = row.count()
        if n_row < 5:
            n = 0
        elif n_row == 5:
            n = 1
        else:
            n = math.ceil(n_row*0.2)
        min = n_row
        row_best = row
        if n == 0:
            return row_best
        for j in range(n+1):
            row_c=row.copy()
            row_c[:j] = np.nan
            row_c[n_row-(n-j):] = np.nan
            if row_c.max()-row_c.min() < min:
                min = row_c.max()-row_c.min()
                row_best = row_c.copy()
        return row_best

    def clean_rows(self, a, d, srr, n, to_save):
        try:
            d_m = np.nanmean(d, axis=0)
        except Exception as err:
            print(err)
        if d.shape[1] > n:
            if d_m[n] < 0.05:
                b = pd.DataFrame(a.copy())
                b.loc[b.apply(lambda x: x.count(), axis=1) > 4, [1]] = np.nan

                # self.save_to_excel(b, srr, 'b '+str(n))
                to_save.append((b, srr, 'b '+str(n)))

                b = b.to_numpy()
                b.sort(axis=1)
                return self.clean_rows(b, d, srr, n+1, to_save)
        return a

    def save_to_excel_(self, to_save, srr):
        with pd.ExcelWriter(srr, engine='openpyxl', mode="a") as writer:
            print(90011)
            for d in to_save:
                print("9006-1")
                print(d[2])
                print("9006-2")
                try:
                    d[0].to_excel(writer, sheet_name=d[2])
                except Exception as ex:
                    print("9006-3 " + d[2] + str(ex))
                print(9007)

    def save_to_excel(self, df, srr, folder):
        df2 = df.copy()
        log_debug(srr + ' -Before sleep save_to_excel- ' + folder)
        # total, used, free = shutil.disk_usage("/")
        # log_debug(' total: ' + str(total) + ' used: ' + str(used) + ' free: ' + str(free))
        nnn = 0
        try:
            with pd.ExcelWriter(srr, mode='a') as writer:
                df2.to_excel(writer, sheet_name=folder)
                writer.save()
                time.sleep(5)
            if self.second_time_save != '':
                print("save ok:", self.second_time_save)
                log_debug("save ok:", self.second_time_save)
            self.second_time_save = ''
            nnn = 1
        except Exception as ee:
            log_debug("exception")
            time.sleep(5)
            self.save_to_excel(df2, srr, folder)
            self.second_time_save = srr
            nnn = 1
        finally:
            log_debug(srr + ' finally -' + str(nnn) + ' - ' + folder)
            if nnn == 0:
                log_debug(srr + ' 55 finally -' + str(nnn) + ' - ' + folder)
                log_debug(srr + ' 55 finally -' + str(nnn) + ' - ' + folder)
                time.sleep(5)
                log_debug(srr + ' 551 finally -' + str(nnn) + ' - ' + folder)
                self.save_to_excel(df2, srr, folder)
                log_debug(srr + ' 66655 finally -' + str(nnn) + ' - ' + folder)
                self.second_time_save = srr

    def get_potential(self, vv, aa_):
        za = sum(aa_)
        z = 0
        for i in range(len(vv)):
            z += (aa_[i]*vv[i])/za
        return z

